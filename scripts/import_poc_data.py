"""
PERKESO POC Data Import Script
Imports data from 'Sample Data POC.xlsx' into Supabase tables.
Usage: python scripts/import_poc_data.py
Requires: pip install openpyxl supabase python-dotenv
"""

import os
import sys
import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# 1. Load environment variables
# ---------------------------------------------------------------------------
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass  # python-dotenv optional; fall back to real env vars

SUPABASE_URL = os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL")
SUPABASE_KEY = (
    os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    or os.environ.get("SERVICE_ROLE_KEY")
)

if not SUPABASE_URL:
    sys.exit("ERROR: SUPABASE_URL not set in .env")
if not SUPABASE_KEY:
    sys.exit(
        "ERROR: SUPABASE_SERVICE_ROLE_KEY not set in .env\n"
        "Add: SUPABASE_SERVICE_ROLE_KEY=<your service role key> to .env"
    )

# ---------------------------------------------------------------------------
# 2. Imports
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl supabase python-dotenv")

try:
    from supabase import create_client, Client
except ImportError:
    sys.exit("ERROR: supabase not installed. Run: pip install openpyxl supabase python-dotenv")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ---------------------------------------------------------------------------
# 3. Helpers
# ---------------------------------------------------------------------------
XLSX_PATH = Path(__file__).parent.parent.parent / "Sample Data POC.xlsx"
if not XLSX_PATH.exists():
    # Also try project root
    alt = Path(__file__).parent.parent / "Sample Data POC.xlsx"
    if alt.exists():
        XLSX_PATH = alt
    else:
        sys.exit(f"ERROR: Could not find 'Sample Data POC.xlsx'. Tried:\n  {XLSX_PATH}\n  {alt}")


def cell_val(cell):
    """Return cell value, converting datetime to ISO string. None for empty."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def to_int(v, default=0):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def batch_insert(table: str, rows: list[dict], batch_size: int = 500):
    """Insert rows in batches, logging progress. Continues on error."""
    total = len(rows)
    inserted = 0
    for start in range(0, total, batch_size):
        chunk = rows[start : start + batch_size]
        try:
            supabase.table(table).insert(chunk).execute()
            inserted += len(chunk)
            print(f"  Imported {inserted}/{total} {table}...")
        except Exception as e:
            print(f"  [ERROR] batch {start}-{start+len(chunk)} for {table}: {e}")
    print(f"  Done: {table} ({inserted}/{total} rows)")


# ---------------------------------------------------------------------------
# 4. Load workbook
# ---------------------------------------------------------------------------
print(f"Loading workbook: {XLSX_PATH}")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
print(f"Sheets found: {wb.sheetnames}")


# ---------------------------------------------------------------------------
# 5. Sheet: JS → poc_candidates
# ---------------------------------------------------------------------------
def import_candidates():
    ws = wb["JS"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    print(f"\n[poc_candidates] Headers: {headers}")

    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        records.append({
            "id":                      get(row, "Candidate"),
            "education_level":         get(row, "Education Level"),
            "nec_1d":                  get(row, "djs_nec1d_name"),
            "nec_2d":                  get(row, "djs_nec2d_name"),
            "nec_3d":                  get(row, "nec3d_list"),
            "institution":             get(row, "Institution"),
            "preferred_occupation":    get(row, "Preferred Name"),
            "preferred_salary":        get(row, "Preferred Salary"),
            "previous_occupation":     get(row, "Previous Occupation"),
            "previous_years_experience": get(row, "Previous Year Experience"),
            "preferred_state":         get(row, "Preferred State"),
            "skills":                  get(row, "Skills"),
        })

    print(f"  Rows to insert: {len(records)}")
    batch_insert("poc_candidates", records, batch_size=500)


# ---------------------------------------------------------------------------
# 6. Sheet: Vac → poc_vacancies
# ---------------------------------------------------------------------------
def import_vacancies():
    ws = wb["Vac"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    print(f"\n[poc_vacancies] Headers: {headers}")

    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        records.append({
            "id":               get(row, "Id"),
            "occupation_name":  get(row, "Occupation Name"),
            "job_title":        get(row, "Job Title"),
            "job_description":  get(row, "Job Description"),
            "education_level":  get(row, "Education Level"),
            "field_of_study":   get(row, "Field Of Study"),
            "state":            get(row, "State"),
            "city":             get(row, "City"),
            "salary":           get(row, "Salary"),
            "skills":           get(row, "Skills"),
        })

    print(f"  Rows to insert: {len(records)}")
    batch_insert("poc_vacancies", records, batch_size=500)


# ---------------------------------------------------------------------------
# 7. Sheet: Vac Manual → poc_vacancies_manual
# ---------------------------------------------------------------------------
def import_vacancies_manual():
    ws = wb["Vac Manual"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    print(f"\n[poc_vacancies_manual] Headers: {headers}")

    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        records.append({
            "employer_code":      get(row, "Emp"),
            "job_title":          get(row, "Job Title"),
            "no_of_positions":    to_int(get(row, "No Of Position")),
            "position_level":     get(row, "Position Level"),
            "education_level":    get(row, "Education Level"),
            "industry":           get(row, "Industry"),
            "salary":             get(row, "Salary"),
            "contract_type":      get(row, "Contract Type"),
            "working_hours":      get(row, "Working Hours"),
            "min_years_experience": to_int(get(row, "Minimum Years Of Working Experiences")),
            "job_description":    get(row, "Job Description"),
        })

    print(f"  Rows to insert: {len(records)}")
    batch_insert("poc_vacancies_manual", records, batch_size=500)


# ---------------------------------------------------------------------------
# 8. Sheet: Behaviour → poc_behaviour
# ---------------------------------------------------------------------------
def import_behaviour():
    ws = wb["Behaviour"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    print(f"\n[poc_behaviour] Headers: {headers}")

    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        records.append({
            "candidate_id":                    get(row, "Candidate"),
            "interview_count":                 to_int(get(row, "Interview")),
            "job_interview_feedback_count":    to_int(get(row, "Job interview feedback")),
            "job_offer_count":                 to_int(get(row, "Job offer")),
            "job_search_count":                to_int(get(row, "Job Search")),
            "report_for_duty_count":           to_int(get(row, "Report for duty")),
            "sign_in_count":                   to_int(get(row, "Sign in")),
            "submitted_application_count":     to_int(get(row, "Submitted application")),
            "grand_total":                     to_int(get(row, "Grand Total")),
        })

    print(f"  Rows to insert: {len(records)}")
    batch_insert("poc_behaviour", records, batch_size=500)


# ---------------------------------------------------------------------------
# 9. Sheet: Log Activity → poc_activity_log
# ---------------------------------------------------------------------------
def import_activity_log():
    ws = wb["Log Activity"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    print(f"\n[poc_activity_log] Headers: {headers}")

    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        records.append({
            "candidate_id":     get(row, "Candidate"),
            "activity_name":    get(row, "Activity Name"),
            "activity_date":    get(row, "Activity Date"),
            "job_title":        get(row, "Job Title"),
            "occupation_name":  get(row, "Occupation Name"),
            "education_level":  get(row, "Education Level"),
            "state":            get(row, "State"),
            "salary":           get(row, "Salary"),
            "vacancy_id":       get(row, "Vacancy ID"),
        })

    print(f"  Rows to insert: {len(records)}")
    batch_insert("poc_activity_log", records, batch_size=1000)


# ---------------------------------------------------------------------------
# 10. Post-import SQL: parse salary and skills
# ---------------------------------------------------------------------------
def run_sql(label: str, sql: str):
    print(f"\n[SQL] {label}")
    try:
        # Use postgrest rpc or direct execute via service role
        result = supabase.rpc("exec_sql", {"query": sql}).execute()
        print(f"  OK: {result}")
    except Exception as e:
        # Fallback: some Supabase setups need direct psql; log and continue
        print(f"  [WARN] rpc exec_sql not available or failed: {e}")
        print(f"  Please run this SQL manually in Supabase SQL editor:\n{sql}\n")


def run_post_import_sql():
    sql_vacancies_salary = """
UPDATE poc_vacancies
SET
  salary_min = CASE
    WHEN salary LIKE '<RM%' THEN 0
    WHEN salary LIKE '>RM%' THEN regexp_replace(salary, '[^0-9]', '', 'g')::int
    ELSE regexp_replace(split_part(salary, '-', 1), '[^0-9]', '', 'g')::int
  END,
  salary_max = CASE
    WHEN salary LIKE '<RM%' THEN regexp_replace(salary, '[^0-9]', '', 'g')::int
    WHEN salary LIKE '>RM%' THEN 99999
    ELSE regexp_replace(split_part(salary, '-', 2), '[^0-9]', '', 'g')::int
  END
WHERE salary IS NOT NULL;
""".strip()

    sql_candidates_salary = """
UPDATE poc_candidates
SET
  salary_min = CASE
    WHEN preferred_salary LIKE '<RM%' THEN 0
    WHEN preferred_salary LIKE '>RM%' THEN regexp_replace(preferred_salary, '[^0-9]', '', 'g')::int
    ELSE regexp_replace(split_part(preferred_salary, '-', 1), '[^0-9]', '', 'g')::int
  END,
  salary_max = CASE
    WHEN preferred_salary LIKE '<RM%' THEN regexp_replace(preferred_salary, '[^0-9]', '', 'g')::int
    WHEN preferred_salary LIKE '>RM%' THEN 99999
    ELSE regexp_replace(split_part(preferred_salary, '-', 2), '[^0-9]', '', 'g')::int
  END
WHERE preferred_salary IS NOT NULL;
""".strip()

    sql_skills = """
UPDATE poc_vacancies
  SET skills_array = to_jsonb(string_to_array(skills, ', '))
  WHERE skills IS NOT NULL AND skills != '';
UPDATE poc_candidates
  SET skills_array = to_jsonb(string_to_array(skills, ', '))
  WHERE skills IS NOT NULL AND skills != '';
""".strip()

    sql_verify = """
SELECT 'poc_vacancies' as t, count(*) FROM poc_vacancies
UNION ALL SELECT 'poc_candidates', count(*) FROM poc_candidates
UNION ALL SELECT 'poc_behaviour', count(*) FROM poc_behaviour
UNION ALL SELECT 'poc_vacancies_manual', count(*) FROM poc_vacancies_manual
UNION ALL SELECT 'poc_activity_log', count(*) FROM poc_activity_log;
""".strip()

    # Print all SQL for manual execution in case rpc is unavailable
    print("\n" + "="*70)
    print("POST-IMPORT SQL (run these in Supabase SQL Editor if not auto-executed):")
    print("="*70)
    print("\n-- 1. Parse vacancy salaries:")
    print(sql_vacancies_salary)
    print("\n-- 2. Parse candidate salaries:")
    print(sql_candidates_salary)
    print("\n-- 3. Parse skills arrays:")
    print(sql_skills)
    print("\n-- 4. Verify counts (expected: 5828, 1449, 1449, 13, 513555):")
    print(sql_verify)
    print("="*70 + "\n")


# ---------------------------------------------------------------------------
# 11. Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import_candidates()
    import_vacancies()
    import_vacancies_manual()
    import_behaviour()
    import_activity_log()
    run_post_import_sql()
    print("\nImport script completed. Check output above for any errors.")
