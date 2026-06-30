"""
Fix/re-run for failed POC imports:
  - poc_vacancies: upsert (duplicate key issue)
  - poc_vacancies_manual: skip trailing empty rows
  - poc_behaviour: skip summary/non-candidate rows (e.g. 'Grand Total')
  - poc_candidates: upsert (in case partial)
Usage: python scripts/fix_poc_import.py
"""

import os
import sys
import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass

SUPABASE_URL = os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    sys.exit("ERROR: SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY not set in .env")

import openpyxl
from supabase import create_client, Client

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

XLSX_PATH = Path(__file__).parent.parent.parent / "Sample Data POC.xlsx"
if not XLSX_PATH.exists():
    alt = Path(__file__).parent.parent / "Sample Data POC.xlsx"
    XLSX_PATH = alt if alt.exists() else None
if not XLSX_PATH:
    sys.exit("ERROR: Sample Data POC.xlsx not found")

print(f"Loading workbook: {XLSX_PATH}")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)


def cell_val(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def to_int(v, default=0):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def batch_upsert(table: str, rows: list[dict], batch_size: int = 500):
    total = len(rows)
    inserted = 0
    for start in range(0, total, batch_size):
        chunk = rows[start: start + batch_size]
        try:
            supabase.table(table).upsert(chunk).execute()
            inserted += len(chunk)
            print(f"  Upserted {inserted}/{total} {table}...")
        except Exception as e:
            print(f"  [ERROR] batch {start}-{start+len(chunk)} for {table}: {e}")
    print(f"  Done: {table} ({inserted}/{total} rows)")


# ── poc_candidates (upsert) ──────────────────────────────────────────────────
def fix_candidates():
    ws = wb["JS"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx].value if hasattr(row[idx], 'value') else row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        cid = cell_val(row[col["Candidate"]].value if hasattr(row[col["Candidate"]], 'value') else row[col["Candidate"]])
        if not cid:
            continue
        records.append({
            "id": cid,
            "education_level": get(row, "Education Level"),
            "nec_1d": get(row, "djs_nec1d_name"),
            "nec_2d": get(row, "djs_nec2d_name"),
            "nec_3d": get(row, "nec3d_list"),
            "institution": get(row, "Institution"),
            "preferred_occupation": get(row, "Preferred Name"),
            "preferred_salary": get(row, "Preferred Salary"),
            "previous_occupation": get(row, "Previous Occupation"),
            "previous_years_experience": get(row, "Previous Year Experience"),
            "preferred_state": get(row, "Preferred State"),
            "skills": get(row, "Skills"),
        })

    print(f"\n[poc_candidates] {len(records)} rows")
    batch_upsert("poc_candidates", records, 500)


# ── poc_vacancies (upsert) ───────────────────────────────────────────────────
def fix_vacancies():
    ws = wb["Vac"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx].value if hasattr(row[idx], 'value') else row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        vid = get(row, "Id")
        if not vid:
            continue
        records.append({
            "id": vid,
            "occupation_name": get(row, "Occupation Name"),
            "job_title": get(row, "Job Title"),
            "job_description": get(row, "Job Description"),
            "education_level": get(row, "Education Level"),
            "field_of_study": get(row, "Field Of Study"),
            "state": get(row, "State"),
            "city": get(row, "City"),
            "salary": get(row, "Salary"),
            "skills": get(row, "Skills"),
        })

    print(f"\n[poc_vacancies] {len(records)} rows")
    batch_upsert("poc_vacancies", records, 500)


# ── poc_vacancies_manual (skip empty rows) ───────────────────────────────────
def fix_vacancies_manual():
    ws = wb["Vac Manual"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx].value if hasattr(row[idx], 'value') else row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        jt = get(row, "Job Title")
        if not jt:
            continue  # skip empty/summary rows
        records.append({
            "employer_code": get(row, "Emp"),
            "job_title": jt,
            "no_of_positions": to_int(get(row, "No Of Position")),
            "position_level": get(row, "Position Level"),
            "education_level": get(row, "Education Level"),
            "industry": get(row, "Industry"),
            "salary": get(row, "Salary"),
            "contract_type": get(row, "Contract Type"),
            "working_hours": get(row, "Working Hours"),
            "min_years_experience": to_int(get(row, "Minimum Years Of Working Experiences")),
            "job_description": get(row, "Job Description"),
        })

    print(f"\n[poc_vacancies_manual] {len(records)} rows")
    batch_upsert("poc_vacancies_manual", records, 500)


# ── poc_behaviour (skip non-candidate rows) ──────────────────────────────────
def fix_behaviour():
    ws = wb["Behaviour"]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    def get(row, name):
        idx = col.get(name)
        return cell_val(row[idx].value if hasattr(row[idx], 'value') else row[idx]) if idx is not None else None

    records = []
    for row in rows_iter:
        cid = get(row, "Candidate")
        if not cid:
            continue
        # Skip summary rows like "Grand Total" — real candidates start with 'C'
        if not str(cid).startswith("C"):
            continue
        records.append({
            "candidate_id": cid,
            "interview_count": to_int(get(row, "Interview")),
            "job_interview_feedback_count": to_int(get(row, "Job interview feedback")),
            "job_offer_count": to_int(get(row, "Job offer")),
            "job_search_count": to_int(get(row, "Job Search")),
            "report_for_duty_count": to_int(get(row, "Report for duty")),
            "sign_in_count": to_int(get(row, "Sign in")),
            "submitted_application_count": to_int(get(row, "Submitted application")),
            "grand_total": to_int(get(row, "Grand Total")),
        })

    print(f"\n[poc_behaviour] {len(records)} rows")
    batch_upsert("poc_behaviour", records, 500)


if __name__ == "__main__":
    fix_candidates()
    fix_vacancies()
    fix_vacancies_manual()
    fix_behaviour()
    print("\nFix import completed.")
    print("\nNow run these SQL statements in Supabase SQL Editor:")
    print("""
-- 1. Parse vacancy salaries:
UPDATE poc_vacancies
SET
  salary_min = CASE
    WHEN salary LIKE '<RM%' THEN 0
    WHEN salary LIKE '>RM%' THEN regexp_replace(salary, '[^0-9]', '', 'g')::int
    ELSE regexp_replace(split_part(salary, '-', 1), '[^0-9]', '', 'g')::int
  END,
  salary_max = CASE
    WHEN salary LIKE '<RM%' THEN regexp_replace(salary, '[^0-9]', '', 'g')::int
    WHEN salary LIKE '>RM%' THEN 99999
    ELSE regexp_replace(split_part(salary, '-', 2), '[^0-9]', '', 'g')::int
  END
WHERE salary IS NOT NULL;

-- 2. Parse candidate salaries:
UPDATE poc_candidates
SET
  salary_min = CASE
    WHEN preferred_salary LIKE '<RM%' THEN 0
    WHEN preferred_salary LIKE '>RM%' THEN regexp_replace(preferred_salary, '[^0-9]', '', 'g')::int
    ELSE regexp_replace(split_part(preferred_salary, '-', 1), '[^0-9]', '', 'g')::int
  END,
  salary_max = CASE
    WHEN preferred_salary LIKE '<RM%' THEN regexp_replace(preferred_salary, '[^0-9]', '', 'g')::int
    WHEN preferred_salary LIKE '>RM%' THEN 99999
    ELSE regexp_replace(split_part(preferred_salary, '-', 2), '[^0-9]', '', 'g')::int
  END
WHERE preferred_salary IS NOT NULL;

-- 3. Parse skills arrays:
UPDATE poc_vacancies SET skills_array = to_jsonb(string_to_array(skills, ', ')) WHERE skills IS NOT NULL AND skills != '';
UPDATE poc_candidates SET skills_array = to_jsonb(string_to_array(skills, ', ')) WHERE skills IS NOT NULL AND skills != '';

-- 4. Verify counts (expected: 5828, 1449, 1449, 13, 513555):
SELECT 'poc_vacancies' as t, count(*) FROM poc_vacancies
UNION ALL SELECT 'poc_candidates', count(*) FROM poc_candidates
UNION ALL SELECT 'poc_behaviour', count(*) FROM poc_behaviour
UNION ALL SELECT 'poc_vacancies_manual', count(*) FROM poc_vacancies_manual
UNION ALL SELECT 'poc_activity_log', count(*) FROM poc_activity_log;
""")
