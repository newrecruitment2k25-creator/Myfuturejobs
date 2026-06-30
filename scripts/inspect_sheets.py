"""Quick inspection of sheet data to understand actual formats."""
import openpyxl
from pathlib import Path

XLSX_PATH = Path(__file__).parent.parent.parent / "Sample Data POC.xlsx"
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

for sheet_name, col_name, n in [
    ("JS", "Candidate", 3),
    ("Behaviour", "Candidate", 5),
    ("Vac Manual", "Job Title", 5),
    ("Vac Manual", "Emp", 5),
]:
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=False)
    headers = [c.value for c in next(rows_iter)]
    col = {h: i for i, h in enumerate(headers) if h is not None}
    idx = col.get(col_name)
    print(f"\n[{sheet_name}] First {n} values of '{col_name}':")
    for i, row in enumerate(rows_iter):
        if i >= n:
            break
        v = row[idx].value if idx is not None else "N/A"
        print(f"  row {i+2}: {repr(v)}")

# Show last 5 rows of Behaviour
ws = wb["Behaviour"]
rows = list(ws.iter_rows(values_only=True))
print(f"\n[Behaviour] Last 5 rows (of {len(rows)-1} data rows):")
for row in rows[-5:]:
    print(f"  {row}")

# Show all rows of Vac Manual
ws = wb["Vac Manual"]
rows = list(ws.iter_rows(values_only=True))
print(f"\n[Vac Manual] All {len(rows)-1} data rows:")
for row in rows:
    print(f"  {row}")
